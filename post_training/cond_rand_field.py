import numpy as np
import util
from tqdm import tqdm
import pandas as pd 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sklearn_crfsuite

class CondRandField:
    @staticmethod
    def evaluate(preds, inputs, masks):
        all_preds_with_crf = []
        
        for pred, input_i, mask in tqdm(zip(preds, inputs, masks), total=len(preds), desc="Processing samples"):
            preds_with_crf = CondRandField.process_sample(pred, input_i, mask)
            all_preds_with_crf.append(preds_with_crf)
    
        return np.array(all_preds_with_crf)

    @staticmethod
    def process_sample(pred, input_i, mask):
        inputs_train, inputs_test = input_i.value
        preds_train, preds_test = pred.value
        labels_train, labels_test = mask.value
        
        database_name, model_name, spectrum_name = pred.database, pred.model, pred.spectrum
        project_name = database_name, model_name, spectrum_name
        
        logits_train = transform_to_two_channel(preds_train)
        logits_test = transform_to_two_channel(preds_test)
        
        preds_after_crf = train_and_predict_with_crf(logits_train, logits_test, 
                                                    inputs_train, inputs_test, 
                                                    labels_train, labels_test)

        evaluate_performance(preds_test, preds_after_crf, labels_test, project_name)
        
        return preds_after_crf
        

def transform_to_two_channel(logits_positives):
    logits_negatives = -logits_positives
    logits = np.concatenate((logits_positives, logits_negatives), axis=1)
    
    return logits
        

def train_and_predict_with_crf(logits_train, logits_test, image_train, image_test, labels_train, labels_test):
    # Extraindo as características das imagens de treinamento e combinando com os logits
    image_features_train = extract_features_from_image(image_train)
    X_train = combine_features(logits_train, image_features_train)
    y_train = [str(label) for label in labels_train.flatten().tolist()]

    # Extraindo as características das imagens de teste e combinando com os logits
    image_features_test = extract_features_from_image(image_test)
    X_test = combine_features(logits_test, image_features_test)
    y_test = [str(label) for label in labels_test.flatten().tolist()]

    # Inicializando o modelo CRF
    crf = sklearn_crfsuite.CRF(
        algorithm='lbfgs',
        max_iterations=100,
        all_possible_transitions=True
    )

    # Ajustar o modelo CRF com os dados de treinamento
    crf.fit([X_train], [y_train])

    # Previsão com o modelo CRF nos dados de teste
    y_pred_crf = crf.predict([X_test])[0]

    # Convertendo as previsões de volta para o formato original
    y_pred_crf_np = np.array([float(pred) for pred in y_pred_crf]).astype(int).reshape(len(labels_test), 1, 
                                                                                       util.IMG_HEIGHT, 
                                                                                       util.IMG_WIDTH)

    return y_pred_crf_np

def extract_features_from_image(image):
    features = []
    num_samples = image.shape[0]
    for sample in range(num_samples):
        for row in range(image.shape[2]):
            for col in range(image.shape[3]):
                pixel = image[sample, 0, row, col]
                feature = {
                    'pixel': pixel,
                    'sample': sample,
                    'row': row,
                    'col': col
                }
                features.append(feature)
    return features

def combine_features(logits, image_features):
    combined_features = []
    for feature in image_features:
        sample = feature['sample']
        row = feature['row']
        col = feature['col']
        combined_feature = feature.copy()
        combined_feature['logit_class_0'] = logits[sample, 0, row, col]
        combined_feature['logit_class_1'] = logits[sample, 1, row, col]
        combined_features.append(combined_feature)
    return combined_features

def evaluate_performance(preds_before_crf, preds_after_crf, labels_test, project_name):
    
    database_name, model_name, spectrum_name = project_name 
    
    preds_before_crf = util.sigmoid(preds_before_crf)
    preds_before_crf = util.thresholding(preds_before_crf)
    
    f1s_before_crf, _ = util.perfomance_metrics(preds_before_crf, labels_test)
    f1s_after_crf, _ = util.perfomance_metrics(preds_after_crf, labels_test)
    
    print(f'F1 Score before CRF: {f1s_before_crf}')
    print(f'F1 Score after CRF: {f1s_after_crf}')

    # Salvar os resultados em um arquivo Excel
    results = {
        'Metric': ['F1 Score before CRF', 'F1 Score after CRF'],
        'Value': [f1s_before_crf, f1s_after_crf]
    }
    df = pd.DataFrame(results)
    
    # Caminho do arquivo Excel
    file_path = 'post_training/f1_scores_crf.xlsx'

    # Salvar os dados no Excel
    df.to_excel(file_path, index=False)

    # Carregar o arquivo Excel para adicionar o título
    book = load_workbook(file_path)
    sheet = book.active

    # Adicionar o título na primeira linha
    title = f"F1 SCORES TABLE: Model: {model_name.upper()}, Spectrum: {spectrum_name.upper()}, Database: {database_name.upper()}"
    sheet.insert_rows(1)
    sheet['A1'] = title

    # Mesclar as células para o título (se necessário)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    # Ajustar largura das colunas
    for col in range(1, len(df.columns) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20

    # Salvar as alterações
    book.save(file_path)
    book.close()
    





