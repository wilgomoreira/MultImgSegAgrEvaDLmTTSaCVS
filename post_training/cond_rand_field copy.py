import torch
import torch.nn as nn
import torch.nn.functional as F
import pydensecrf.densecrf as dcrf
from pydensecrf.utils import unary_from_softmax, create_pairwise_bilateral, create_pairwise_gaussian
import numpy as np
import util
from tqdm import tqdm
import pandas as pd 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
from scipy.stats import gaussian_kde
from sklearn.mixture import GaussianMixture

device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')

class ConvNet(nn.Module):
    def __init__(self, in_channels=3):
        super(ConvNet, self).__init__()
        self.conv1 = nn.Conv2d(in_channels, 16, kernel_size=3, stride=2, padding=1)
        self.conv2 = nn.Conv2d(16, 32, kernel_size=3, stride=2, padding=1)
        self.conv3 = nn.Conv2d(32, 64, kernel_size=3, stride=2, padding=1)
        self.dropout = nn.Dropout(0.5)

    def forward(self, x):
        x = F.relu(self.conv1(x))
        x = F.relu(self.conv2(x))
        x = self.dropout(x)
        x = F.relu(self.conv3(x))
        x = self.dropout(x)
        return x

# Initialize ConvNet instances for different input channels
convnet_rgb = ConvNet(in_channels=3).to(device)
convnet_grayscale = ConvNet(in_channels=1).to(device)
convnet_early_fusion = ConvNet(in_channels=5).to(device)

class CondRandField:
    @staticmethod
    def evaluate(preds, inputs, masks):
        all_preds_no_crf_train, all_preds_crf_train, all_preds_crf_cnn_train, all_labels_train = [], [], [], []
        all_preds_no_crf_test, all_preds_crf_test, all_preds_crf_cnn_test, all_labels_test = [], [], [], []
        
        for pred, input_i, mask in tqdm(zip(preds, inputs, masks), total=len(preds), desc="Processing samples"):
            (preds_no_crf_train, preds_crf_train, preds_crf_cnn_train, labels_train,
             preds_no_crf_test, preds_crf_test, preds_crf_cnn_test, labels_test) = CondRandField.process_sample(pred, input_i, mask)
            
            all_preds_no_crf_train.append(preds_no_crf_train)
            all_preds_crf_train.append(preds_crf_train)
            all_preds_crf_cnn_train.append(preds_crf_cnn_train)
            all_labels_train.append(labels_train)
            
            all_preds_no_crf_test.append(preds_no_crf_test)
            all_preds_crf_test.append(preds_crf_test)
            all_preds_crf_cnn_test.append(preds_crf_cnn_test)
            all_labels_test.append(labels_test)
        
        return (np.array(all_preds_no_crf_train), np.array(all_preds_crf_train), np.array(all_preds_crf_cnn_train), np.array(all_labels_train),
                np.array(all_preds_no_crf_test), np.array(all_preds_crf_test), np.array(all_preds_crf_cnn_test), np.array(all_labels_test))

    @staticmethod
    def process_sample(pred, input_i, mask):
        inputs_train, inputs_test = input_i.value
        preds_train, preds_test = pred.value
        labels_train, labels_test = mask.value
        
        # Check similarity between training and test data
  #      if not check_similarity(preds_train, preds_test):
       #     raise ValueError("Training and test data are too different, KDE might not be appropriate.")
        
        database_name, model_name, spectrum_name = pred.database, pred.model, pred.spectrum
        project_name = database_name, model_name, spectrum_name
         
        preds_no_crf_train, preds_crf_train, preds_crf_cnn_train = [], [], []
        preds_no_crf_test, preds_crf_test, preds_crf_cnn_test = [], [], []
        
        for img_train, pred_train in tqdm(zip(inputs_train, preds_train), total=len(preds_train), desc="Processing train images"):
            
            pred_train = util.sigmoid(pred_train)
            
            # Select the appropriate convnet based on the number of channels
            if img_train.shape[0] == 3:
                convnet = convnet_rgb
            elif img_train.shape[0] == 1:
                convnet = convnet_grayscale
            elif img_train.shape[0] == 5:
                convnet = convnet_early_fusion
            else:
                raise ValueError(f"Unexpected number of channels: {img_train.shape[0]}")
            
            preds_no_crf_train.append(pred_train)
            preds_crf_train.append(CondRandField.apply_crf(pred_train, img_train, use_cnn=False))
            preds_crf_cnn_train.append(CondRandField.apply_crf(pred_train, img_train, use_cnn=True, convnet=convnet))
            
        
        for img_test, pred_test in tqdm(zip(inputs_test, preds_test), total=len(inputs_test), desc="Processing test images"):
           
            if util.NOT_SOFTMAX:
                pred_test = apply_gmm(pred_test, preds_train)
            else:  
                pred_test = util.sigmoid(pred_test)
            
            # Select the appropriate convnet based on the number of channels
            if img_test.shape[0] == 3:
                convnet = convnet_rgb
            elif img_test.shape[0] == 1:
                convnet = convnet_grayscale
            elif img_test.shape[0] == 5:
                convnet = convnet_early_fusion
            else:
                raise ValueError(f"Unexpected number of channels: {img_test.shape[0]}")
            
            preds_no_crf_test.append(pred_test)
            preds_crf_test.append(CondRandField.apply_crf(pred_test, img_test, use_cnn=False))
            preds_crf_cnn_test.append(CondRandField.apply_crf(pred_test, img_test, use_cnn=True, convnet=convnet))
        
        preds_no_crf_train = np.array(preds_no_crf_train)
        preds_crf_train = np.array(preds_crf_train)
        preds_crf_cnn_train = np.array(preds_crf_cnn_train)
        labels_train = np.array(labels_train)

        preds_no_crf_test = np.array(preds_no_crf_test)
        preds_crf_test = np.array(preds_crf_test)
        preds_crf_cnn_test = np.array(preds_crf_cnn_test)
        labels_test = np.array(labels_test)
        
        evaluate_performance(preds_no_crf_train, preds_crf_train, preds_crf_cnn_train, labels_train,
                          preds_no_crf_test, preds_crf_test, preds_crf_cnn_test, labels_test, project_name)
        
        return (preds_no_crf_train, preds_crf_train, preds_crf_cnn_train, labels_train,
                preds_no_crf_test, preds_crf_test, preds_crf_cnn_test, labels_test)

    @staticmethod
    def apply_crf(pred, image, use_cnn=False, convnet=None):
        h, w = pred.shape[1:]
        
        back_preds = 1 - pred
        fore_preds = pred
        pred = np.vstack((back_preds, fore_preds)).reshape((2, util.IMG_HEIGHT, 
                                                            util.IMG_WIDTH))
        
        unary = unary_from_softmax(pred) 
        d = dcrf.DenseCRF2D(w, h, 2)
        d.setUnaryEnergy(unary)
       
        if use_cnn and convnet is not None:
            image = CondRandField.preprocess_image(image)
        else:
            image = CondRandField.preprocess_image(image)
        
        if image.dtype != np.uint8:
            image = (image * 255).astype(np.uint8)
        else:
            image = image
            
        if (image.shape != (240, 240, 3)) and (image.shape != (240, 240, 64)):
            image = np.transpose(image, (1, 2, 0))
            
        if not image.flags['C_CONTIGUOUS']:
            image = np.ascontiguousarray(image)
            
        pairwise_energy_gaussian = create_pairwise_gaussian(sdims=(3, 3), shape=image.shape)
        d.addPairwiseEnergy(pairwise_energy_gaussian, compat=3)
        
        pairwise_energy_bilateral = create_pairwise_bilateral(sdims=(80, 80),
                                                              schan=(13, 13, 13, 13, 13), 
                                                              img=image, chdim=2)

        d.addPairwiseEnergy(pairwise_energy_bilateral, compat=10)
        
        
        Q = d.inference(util.INFERENCE)
        Q_np = np.array(Q)
        Q_reshaped = Q_np.reshape((-1, util.IMG_HEIGHT, util.IMG_WIDTH))
        probs = Q_reshaped[1, :, :]
        new_prob = np.expand_dims(probs, axis=0)
        return new_prob

    @staticmethod
    def preprocess_image(image):
        if image.shape[0] != 3 and image.shape[0] != 1:
            image = CondRandField.normalise_enphasise_colours(image)
            image = np.squeeze(image, axis=0)
        return np.ascontiguousarray(image, dtype=np.uint8)

    @staticmethod
    def process_image_with_cnn(image, convnet):
        image_tensor = torch.tensor(image, dtype=torch.float32)

        # Ensure the image tensor has shape (channels, height, width)
        if len(image_tensor.shape) == 2:
            image_tensor = image_tensor.unsqueeze(0)  # Add channel dimension
        elif len(image_tensor.shape) == 3:
            if image_tensor.shape[0] in [1, 3, 5]:
                pass
            elif image_tensor.shape[-1] in [1, 3, 5]:
                image_tensor = image_tensor.permute(2, 0, 1)  # Move channel dimension to the first position
            else:
                raise ValueError(f"Unexpected channel dimension: {image_tensor.shape}")

        image_tensor = image_tensor.unsqueeze(0)  # Add batch dimension
        
        with torch.no_grad():
            features = convnet(image_tensor.to(device)).squeeze(0).permute(1, 2, 0).cpu().numpy()
        
        h, w = image.shape[1:]
        features = F.interpolate(torch.tensor(features).unsqueeze(0).permute(0, 3, 1, 2), 
                                size=(h, w), mode='bilinear', align_corners=False).squeeze(0)
        return features.permute(1, 2, 0).numpy().astype(np.uint8)
    
    @staticmethod
    def normalise_enphasise_colours(image):
        if image.shape[0] == 3:
            return image
        else:
            image = (image + 1) / 2  # Normaliza para o intervalo [0, 1]
            colors = [(165/255, 42/255, 42/255), (255/255, 255/255, 224/255), (34/255, 139/255, 34/255)]
            n_bins = [0, 0.5, 1]
            cmap_name = 'my_list'
            cm = mcolors.LinearSegmentedColormap.from_list(cmap_name, list(zip(n_bins, colors)))
            image_colored = cm(image)  # Aplica a paleta de cores
            return image_colored[..., :3]  # Remove o canal alpha


@staticmethod
def evaluate_performance(preds_no_crf_train, preds_crf_train, preds_crf_cnn_train, labels_train,
                    preds_no_crf_test, preds_crf_test, preds_crf_cnn_test, labels_test, project_name):
    
    database_name, model_name, spectrum_name = project_name 
    
   
    preds_no_crf_train = util.thresholding(preds_no_crf_train)
    preds_crf_train = util.thresholding(preds_crf_train)
    preds_crf_cnn_train = util.thresholding(preds_crf_cnn_train)
    
    f1s_no_crf_train, _ = util.perfomance_metrics(preds_no_crf_train, labels_train)
    f1s_crf_train, _ = util.perfomance_metrics(preds_crf_train, labels_train)
    f1s_crf_cnn_train, _ = util.perfomance_metrics(preds_crf_cnn_train, labels_train)

    print(f'F1 Score without CRF (train): {f1s_no_crf_train}')
    print(f'F1 Score with CRF (train): {f1s_crf_train}')
    print(f'F1 Score with CRF and CNN (train): {f1s_crf_cnn_train}')

    
    preds_no_crf_test = util.thresholding(preds_no_crf_test)
    preds_crf_test = util.thresholding(preds_crf_test)
    preds_crf_cnn_test = util.thresholding(preds_crf_cnn_test)
    
    f1s_no_crf_test, _ = util.perfomance_metrics(preds_no_crf_test, labels_test)
    f1s_crf_test, _ = util.perfomance_metrics(preds_crf_test, labels_test)
    f1s_crf_cnn_test, _ = util.perfomance_metrics(preds_crf_cnn_test, labels_test)
  
    print(f'F1 Score without CRF (test): {f1s_no_crf_test}')
    print(f'F1 Score with CRF (test): {f1s_crf_test}')
    print(f'F1 Score with CRF and CNN (test): {f1s_crf_cnn_test}')
    
    # Salvar os resultados em um arquivo Excel
    results = {
        'Metric': ['F1 Score without CRF', 'F1 Score with CRF', 'F1 Score with CRF and CNN'],
        'Train': [f1s_no_crf_train, f1s_crf_train, f1s_crf_cnn_train],
        'Test': [f1s_no_crf_test, f1s_crf_test, f1s_crf_cnn_test]
    }
    df = pd.DataFrame(results)
    
    # Caminho do arquivo Excel
    file_path = 'post_training/f1_scores.xlsx'

    # Salvar os dados no Excel
    df.to_excel(file_path, index=False)

    # Carregar o arquivo Excel para adicionar o título
    book = load_workbook(file_path)
    sheet = book.active

    # Adicionar o título na primeira linha
    title = f"F1 SCORES TABLE: Model: {model_name.upper()}, Spectrum: {spectrum_name.upper()}, Database: {database_name.upper()}"
    sheet.insert_rows(1)
    sheet['A1'] = title

    # Mesclar as células para o título (se necessário)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))

    # Ajustar largura das colunas
    for col in range(1, len(df.columns) + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20

    # Salvar as alterações
    book.save(file_path)
    book.close()
    

@staticmethod
def apply_kde(pred, preds_train):
    # Flatten the prediction and training arrays
    pred_flat = pred.flatten()
    preds_train_flat = preds_train.flatten()
    
    # Create KDE based on the training data
    kde = gaussian_kde(preds_train_flat)
    
    # Evaluate the KDE on the flattened prediction
    likelihoods = kde.evaluate(pred_flat)
    
    # Normalize the likelihoods
    normalized_likelihoods = likelihoods / np.sum(likelihoods)
    
    # Reshape to the original prediction shape
    normalized_likelihoods = normalized_likelihoods.reshape(pred.shape)
    
    return normalized_likelihoods

@staticmethod
def check_similarity(train_data, test_data, threshold=0.5):
    mean_diff = np.mean(np.abs(np.mean(train_data) - np.mean(test_data)))
    std_diff = np.mean(np.abs(np.std(train_data) - np.std(test_data)))
    similarity_score = mean_diff + std_diff
    return similarity_score < threshold


@staticmethod
def apply_gmm(pred, preds_train, n_components=3):
    # Check dimensions
    n_samples_train, n_channels_train, height_train, width_train = preds_train.shape
    n_channels_pred, height_pred, width_pred = pred.shape

    if n_channels_train != n_channels_pred or height_train != height_pred or width_train != width_pred:
        raise ValueError("The dimensions of pred and preds_train must match in terms of channels, height, and width")
    
    # Flatten the prediction array
    pred_flat = pred.reshape(n_channels_pred, -1)
    
    # Initialize array to store likelihoods
    likelihoods = np.zeros_like(pred_flat)
    
    # Calculate likelihoods for each channel separately
    for i in range(n_channels_pred):
        # Extract the data for the current channel
        preds_train_channel = preds_train[:, i, :, :].reshape(-1, 1)
        pred_channel = pred_flat[i, :].reshape(-1, 1)

        # Fit a Gaussian Mixture Model to the training data for this channel
        gmm = GaussianMixture(n_components=n_components, covariance_type='full')
        gmm.fit(preds_train_channel)
        
        # Evaluate the GMM on the flattened prediction
        likelihoods[i, :] = gmm.score_samples(pred_channel)
    
    # Normalize the likelihoods
    normalized_likelihoods = np.exp(likelihoods) / np.sum(np.exp(likelihoods), axis=1, keepdims=True)
    
    # Reshape to the original prediction shape
    normalized_likelihoods = normalized_likelihoods.reshape(pred.shape)
    
    return normalized_likelihoods


      
def _cond_rand_field(inputs, preds, masks):
    # Evaluate CRF with training and test data
    (preds_no_crf_train, preds_crf_train, preds_crf_cnn_train, labels_train,
     preds_no_crf_test, preds_crf_test, 
     preds_crf_cnn_test, labels_test) = CondRandField.evaluate(preds=preds, 
                                                               inputs=inputs, 
                                                               masks=masks)

    return preds_no_crf_train, preds_crf_train, preds_crf_cnn_train, preds_no_crf_test, preds_crf_test, preds_crf_cnn_test





